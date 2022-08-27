import numpy as np
import matplotlib.pyplot as plt
from matplotlib import animation

from openpyxl import load_workbook


def get_table(ws):
    """
    get the data on a worksheet as a numpy array
    """
    ret = []

    for v in ws.iter_rows():
        if v is None:
            break

        # openpyxl doesn't support calulating the formula, doesn't matter
        ret.append(list(map(lambda c: c.value, v)))

    return np.array(ret)


def retrieve_data(filename):
    """
    retrieve X and 1/-r from an excel workbook
    """
    wb = load_workbook(filename)
    table = get_table(wb.active)

    table = table.transpose()

    # skip the header and remove the last points
    X = np.array(table[0][1:-2], dtype=float)
    r = np.array(table[1][1:-2], dtype=float)

    return X, r


def closest_index(l, v):
    for i, vi in enumerate(l[:-1]):
        if vi == v:
            return i, i

        if v > vi and v < l[i + 1]:
            return i, i + 1

    return None


X, r = retrieve_data("dataf.xlsx")

fig = plt.figure()

ax = fig.add_subplot(111)
ax.set_xlabel("$X$")
ax.set_ylabel("$\\frac{1}{-r_DCBO}$")
ax.set_xticks(np.arange(0, 1, step=0.1))
ax.set_xlim([0, 1])
ax.set_ylim([0, r[-1] * 1.1])

ax.plot(X, r)

last_i = closest_index(X, 0.8)
assert last_i is not None

def abserr(m, r):
    return abs((m - r) / r) * 100

# kinda hacky, checkout later https://matplotlib.org/stable/gallery/animation/pause_resume.html
anim_done = False


# https://stackoverflow.com/questions/31453422/displaying-numbers-with-x-instead-of-e-scientific-notation-in-matplotlib
def as_si(x, ndp):
    s = '{x:0.{ndp:d}e}'.format(x=x, ndp=ndp)
    m, e = s.split('e')
    return r'{m:s}\times 10^{{{e:d}}}'.format(m=m, e=int(e))

total_A_text = ax.text(0.7, 1e7, r"$A = {0:s}$".format(as_si(0, 2)),
    {"horizontalalignment": "center",
     "verticalalignment": "center"})

pfr_A_text = ax.text(0, 1e7, r"$A = {0:s}$".format(as_si(0, 2)),
    {"horizontalalignment": "center",
     "verticalalignment": "center"})

cstr_A_text = ax.text(0, 1e7, r"$A = {0:s}$".format(as_si(0, 2)),
    {"horizontalalignment": "center",
     "verticalalignment": "center"})


def pfr_cstr_animate(i):
    global anim_done

    if anim_done:
        return

    if i > 41:
        i = 42 - (i - 42) - 1

    pfr_X = X[:i]
    pfr_r = r[:i]

    ax.collections.clear()
    
    ax.fill_between(pfr_X, pfr_r, facecolor="yellow")

    cstr_X = []
    cstr_r = [r[last_i[1]], r[last_i[1]]]

    if len(pfr_X):
        cstr_X = [pfr_X[-1], 0.8]
    else:
        cstr_X = [0, 0.8]

    ax.fill_between(cstr_X, cstr_r, facecolor="magenta")

    pfr_A = np.trapz(pfr_r, pfr_X)
    cstr_A = cstr_r[-1] * (cstr_X[-1] - cstr_X[0])
    total_A = pfr_A + cstr_A

    if abserr(pfr_A, cstr_A) < 5.0:
        anim_done = True

    print(pfr_A, cstr_A)

    total_A_text.set_text(r"$A = {0:s}$".format(as_si(total_A, 2)))

    pfr_A_text.set_text(r"$A = {0:s}$".format(as_si(pfr_A, 2)))

    if len(pfr_X):
        pfr_A_text.set_position(((pfr_X[0] + pfr_X[-1]) / 2, 1e7 / 12))
    else:
        pfr_A_text.set_position((0, 1e7/12))

    cstr_A_text.set_text(r"$A = {0:s}$".format(as_si(cstr_A, 2)))
    cstr_A_text.set_position(((cstr_X[0] + cstr_X[-1]) / 2, 1e7 / 7))

    # return fill,


a = animation.FuncAnimation(fig, pfr_cstr_animate, frames=42 * 2 - 1, interval=100)
a.save("anim.m4a")


plt.show()
